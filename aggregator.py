# aggregator.py — merges all normalized DataFrames and writes Excel + JSON output

import os
import re
import glob
import json
import hashlib
import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import normalizer
from config import OUTPUT_COLUMNS, MANUAL_SOURCES


OUTPUT_DIR    = os.path.join(os.path.dirname(__file__), "output")
IMPORTS_DIR   = os.path.join(os.path.dirname(__file__), "imports")
DOCS_DATA_DIR = os.path.join(os.path.dirname(__file__), "docs", "data")


# ---------------------------------------------------------------------------
# Manual import loader
# ---------------------------------------------------------------------------

def _source_name_from_filename(filename: str) -> str:
    base = os.path.splitext(os.path.basename(filename))[0].lower()
    for key, display_name in MANUAL_SOURCES.items():
        if key in base:
            return display_name
    return base.replace("_", " ").replace("-", " ").title()


def _source_from_sheet(sheet_name: str) -> str:
    """Resolve a sheet/tab name to a vendor display name.
    Checks MANUAL_SOURCES keys first for canonical names, then returns
    the sheet name as-is (it IS the vendor label in multi-tab files).
    """
    lower = sheet_name.strip().lower()
    for key, display_name in MANUAL_SOURCES.items():
        if key in lower:
            return display_name
    return sheet_name.strip()


def _file_hash(filepath: str) -> str:
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        h.update(f.read())
    return h.hexdigest()


def load_manual_imports() -> list[pd.DataFrame]:
    """Load all CSV/Excel files from the imports/ directory, skipping duplicates."""
    frames = []
    patterns = ["*.xlsx", "*.xls", "*.csv"]
    files = []
    for pat in patterns:
        files.extend(glob.glob(os.path.join(IMPORTS_DIR, pat)))

    seen_hashes: dict[str, str] = {}

    for filepath in sorted(files):
        basename = os.path.basename(filepath)
        digest = _file_hash(filepath)
        if digest in seen_hashes:
            print(f"  [import] SKIPPED {basename} — duplicate of '{seen_hashes[digest]}'")
            continue
        seen_hashes[digest] = basename

        source_name = _source_name_from_filename(filepath)
        try:
            if filepath.endswith(".csv"):
                raw = pd.read_csv(filepath, dtype=str)
                raw["_raw_source"] = source_name
                df = normalizer.normalize(raw, source_name)
                if not df.empty:
                    frames.append(df)
                    print(f"  [import] {source_name}: {len(df)} rows from {basename}")
            else:
                # Read all sheets — multi-tab vendor files are common
                sheets = pd.read_excel(filepath, sheet_name=None, dtype=str)
                for sheet_name, raw in sheets.items():
                    if raw.empty:
                        continue
                    # Tab name IS the vendor for multi-sheet files
                    sheet_source = _source_from_sheet(sheet_name)
                    raw["_raw_source"] = sheet_source
                    df = normalizer.normalize(raw, sheet_source)
                    if not df.empty:
                        frames.append(df)
                        print(f"  [import] {sheet_source}: {len(df)} rows from {basename} [{sheet_name}]")
        except Exception as exc:
            print(f"  [import] ERROR loading {filepath}: {exc}")

    return frames


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)


def _style_sheet(ws, freeze_row: int = 1):
    for cell in ws[freeze_row]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(ws.iter_rows(min_row=freeze_row + 1), start=1):
        if i % 2 == 0:
            for cell in row:
                cell.fill = _ALT_ROW_FILL
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
    ws.freeze_panes = ws.cell(row=freeze_row + 1, column=1)


# ---------------------------------------------------------------------------
# Main aggregation + write
# ---------------------------------------------------------------------------

def write_excel(frames: list[pd.DataFrame]) -> str:
    """
    Combine all DataFrames, write an Excel file with Master + per-source sheets.
    Returns the output file path.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = datetime.date.today().strftime("%Y-%m-%d")
    out_path = os.path.join(OUTPUT_DIR, f"inventory_{today}.xlsx")

    if not frames:
        print("  [aggregator] No data to write.")
        return ""

    master = pd.concat(frames, ignore_index=True)

    # Dedup by serial AND inv — both are unique identifiers for a physical machine.
    # Priority: serial > inv number > no identifier (keep all).
    def _score_df(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df["_score"] = (
            (pd.to_numeric(df.get("price",       pd.Series(dtype=float)), errors="coerce").fillna(0) > 0).astype(int) * 4 +
            (pd.to_numeric(df.get("total_meter", pd.Series(dtype=float)), errors="coerce").fillna(0) > 0).astype(int) * 2 +
            (df.get("description", pd.Series(dtype=str)).fillna("").str.len() > 0).astype(int)
        )
        return df

    if "serial" in master.columns:
        has_serial = master["serial"].notna() & (master["serial"].astype(str).str.strip() != "")
    else:
        has_serial = pd.Series(False, index=master.index)

    if "inv" in master.columns:
        has_inv = master["inv"].notna() & (master["inv"].astype(str).str.strip() != "")
    else:
        has_inv = pd.Series(False, index=master.index)

    # Group 1: has serial → dedup by serial
    grp_serial = master[has_serial].copy()
    # Group 2: no serial, has inv → dedup by inv
    grp_inv    = master[~has_serial & has_inv].copy()
    # Group 3: no serial, no inv → keep all
    grp_none   = master[~has_serial & ~has_inv].copy()

    if not grp_serial.empty:
        grp_serial = _score_df(grp_serial)
        grp_serial = grp_serial.sort_values("_score", ascending=False)
        grp_serial = grp_serial.drop_duplicates(subset=["serial"], keep="first")
        grp_serial = grp_serial.drop(columns=["_score"])

    if not grp_inv.empty:
        grp_inv = _score_df(grp_inv)
        grp_inv = grp_inv.sort_values("_score", ascending=False)
        grp_inv = grp_inv.drop_duplicates(subset=["inv"], keep="first")
        grp_inv = grp_inv.drop(columns=["_score"])

    master = pd.concat([grp_serial, grp_inv, grp_none], ignore_index=True)

    master = master.sort_values(["source", "brand", "model"]).reset_index(drop=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Master", index=False)
        _style_sheet(writer.sheets["Master"])

        for source in master["source"].unique():
            sub = master[master["source"] == source].reset_index(drop=True)
            # Sanitize: Excel forbids \ / * ? : [ ] in sheet names
            sheet_name = re.sub(r'[\\/*?:\[\]]', '_', source)[:31]
            sub.to_excel(writer, sheet_name=sheet_name, index=False)
            # openpyxl may rename the sheet (e.g. case-conflict → "Copex1");
            # always reference the last-added sheet rather than by name.
            _style_sheet(list(writer.sheets.values())[-1])

    print(f"\n  Output written to: {out_path}")
    print(f"  Total rows: {len(master)}")

    _write_json(master)

    return out_path


# ---------------------------------------------------------------------------
# JSON output for the web UI
# ---------------------------------------------------------------------------

# Mapping from Python column names → JSON field names (matching HTML reference)
_JSON_FIELD_MAP = {
    "source":       "vendor",
    "brand":        "brand",
    "model":        "model",
    "condition":    "condition",
    "state":        "state",
    "inv":          "inv",
    "serial":       "serial",
    "total_meter":  "total",
    "color_meter":  "color",
    "bw_meter":     "bw",
    "is_color":     "isColor",
    "feeder_model": "feederModel",
    "capacity":     "capacity",
    "finisher":     "finisher",
    "print_speed":  "print",
    "scan":         "scan",
    "fax":          "fax",
    "qty":          "qty",
    "price":        "price",
    "description":  "description",
    "notes":        "notes",
}


def _build_config(rec: dict) -> str:
    """
    Build the pipe-separated config string shown in the ⚙ Configuration column.
    E.g.: "RADF | 3x550+LCT | Booklet Finisher | Print | Scan | Fax"
    """
    parts = []
    feeder = rec.get("feederModel", "")
    if feeder:
        parts.append(feeder)
    cap = rec.get("capacity", "")
    if cap:
        parts.append(cap)
    fin = rec.get("finisher", "")
    if fin:
        parts.append(fin)
    if rec.get("print") in ("YES", "yes"):
        parts.append("Print")
    if rec.get("scan") in ("YES", "yes"):
        parts.append("Scan")
    if rec.get("fax") in ("YES", "yes"):
        parts.append("Fax")
    return " | ".join(parts)


def _load_prev_ids() -> tuple[set, set, str]:
    """
    Load the previous inventory.json and return:
      (set of serials, set of inv numbers, updated timestamp)
    Both sets are used for isNew detection.
    """
    json_path = os.path.join(DOCS_DATA_DIR, "inventory.json")
    if not os.path.exists(json_path):
        return set(), set(), ""
    try:
        with open(json_path, encoding="utf-8") as f:
            prev = json.load(f)
        serials = {r.get("serial", "") for r in prev.get("records", []) if r.get("serial", "")}
        invs    = {r.get("inv",    "") for r in prev.get("records", []) if r.get("inv",    "")}
        return serials, invs, prev.get("updated", "")
    except Exception:
        return set(), set(), ""


def _write_json(master: pd.DataFrame):
    """Write docs/data/inventory.json for the GitHub Pages UI."""
    os.makedirs(DOCS_DATA_DIR, exist_ok=True)
    json_path = os.path.join(DOCS_DATA_DIR, "inventory.json")

    # Load previous inventory to detect new items
    prev_serials, prev_invs, prev_updated = _load_prev_ids()

    # Clean the DataFrame
    clean = master.where(master.notna(), other=None).copy()

    # Coerce numeric columns
    for col in ("qty", "price", "total_meter", "color_meter", "bw_meter"):
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce")

    now_iso = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    records = []

    for _, row in clean.iterrows():
        rec: dict = {}
        for py_col, js_field in _JSON_FIELD_MAP.items():
            val = row.get(py_col)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                # Use 0 for numeric fields, "" for strings
                val = 0 if js_field in ("total", "color", "bw", "qty", "price") else ""
            elif isinstance(val, float):
                # Convert whole-number floats to int for cleaner JSON
                val = int(val) if val == int(val) else val
            rec[js_field] = val

        # Determine isNew — new if neither serial nor inv was seen before
        serial = rec.get("serial", "")
        inv_id = rec.get("inv", "")
        serial_new = bool(serial and serial not in prev_serials)
        inv_new    = bool(inv_id and inv_id not in prev_invs)
        # If we have a serial, use it; if only inv, use that; if neither, treat as new
        if serial:
            rec["isNew"] = serial_new
        elif inv_id:
            rec["isNew"] = inv_new
        else:
            rec["isNew"] = False

        # Build config string
        rec["config"] = _build_config(rec)

        records.append(rec)

    # Collect unique states for filter dropdowns
    states = sorted({r.get("state", "") for r in records if r.get("state", "")})

    new_count = sum(1 for r in records if r.get("isNew"))

    payload = {
        "updated":  now_iso,
        "newSince": prev_updated,
        "total":    len(records),
        "newCount": new_count,
        "sources":  sorted(master["source"].unique().tolist()),
        "brands":   sorted(master["brand"].dropna().replace("", pd.NA).dropna().unique().tolist()),
        "states":   states,
        "records":  records,
    }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    print(f"  JSON written to:   {json_path}")
    print(f"  New items:         {new_count}")

    # Also check watchlist matches and write notifications file
    _write_watchlist_matches(records)


def _write_watchlist_matches(records: list[dict]):
    """
    Read docs/data/watchlist.json (if it exists), find matches in current
    inventory, and write docs/data/watchlist_matches.json for the UI.
    """
    watchlist_path = os.path.join(DOCS_DATA_DIR, "watchlist.json")
    if not os.path.exists(watchlist_path):
        return

    try:
        with open(watchlist_path, encoding="utf-8") as f:
            watchlist = json.load(f)
    except Exception:
        return

    if not isinstance(watchlist, list) or not watchlist:
        return

    new_items = [r for r in records if r.get("isNew")]

    match_report = []
    for req in watchlist:
        new_matches = _find_matches(req, new_items)
        if new_matches:
            match_report.append({
                "customerId":  req.get("id"),
                "customerName": req.get("name"),
                "email":       req.get("email", ""),
                "phone":       req.get("phone", ""),
                "matches":     new_matches[:20],
                "matchCount":  len(new_matches),
                "criteria":    {k: req.get(k) for k in ("brand", "model", "maxMeter", "maxPrice", "color", "state", "finisher", "fax")},
            })

    matches_path = os.path.join(DOCS_DATA_DIR, "watchlist_matches.json")
    with open(matches_path, "w", encoding="utf-8") as f:
        json.dump({
            "checkedAt": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "totalNewMatches": sum(r["matchCount"] for r in match_report),
            "requests": match_report,
        }, f, ensure_ascii=False, separators=(",", ":"))

    if match_report:
        total = sum(r["matchCount"] for r in match_report)
        print(f"  Watchlist matches: {total} new matches across {len(match_report)} customer request(s)")
        print(f"  Matches file:      {matches_path}")


def _find_matches(req: dict, inventory: list[dict]) -> list[dict]:
    """Check a watchlist request against a list of inventory records."""
    results = []
    brand    = (req.get("brand") or "").strip().lower()
    model    = (req.get("model") or "").strip().lower()
    color    = req.get("color") or ""
    state    = (req.get("state") or "").strip().upper()
    finisher = req.get("finisher") or ""
    fax      = req.get("fax") or ""
    max_meter = _safe_num(req.get("maxMeter"))
    max_price = _safe_num(req.get("maxPrice"))

    for r in inventory:
        if brand and r.get("brand", "").lower() != brand:
            continue
        if model:
            models = [m.strip() for m in model.split(",") if m.strip()]
            inv_model = r.get("model", "").lower().strip()
            if not inv_model or not any(m in inv_model for m in models):
                continue
        if color and color.lower() != "any" and r.get("isColor", "") != color:
            continue
        if state and r.get("state", "").upper() != state:
            continue
        if finisher == "yes" and not r.get("finisher"):
            continue
        if fax == "YES" and r.get("fax") != "YES":
            continue
        if max_meter is not None:
            total = _safe_num(r.get("total")) or 0
            if total > max_meter:
                continue
        if max_price is not None:
            price = _safe_num(r.get("price"))
            if price is None or price > max_price:
                continue
        results.append(r)

    return results


def _safe_num(val) -> float | None:
    try:
        v = float(str(val or "").replace(",", ""))
        return v if v >= 0 else None
    except (ValueError, TypeError):
        return None
